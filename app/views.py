from flask import Blueprint, render_template,session, redirect, url_for, flash, request, jsonify, current_app, get_flashed_messages
from flask_login import login_user, logout_user, current_user, login_required
from werkzeug.security import generate_password_hash, check_password_hash
from .models import Comment
from werkzeug.utils import secure_filename
from itsdangerous import URLSafeSerializer, URLSafeTimedSerializer, SignatureExpired, BadSignature
import os
from sqlalchemy import  func
from .forms import LoginForm, RegistrationForm
from .models import User, Product, Category, Location, Project, ProductManager, Loan
from .extensions import db, upload_dir
import csv
from datetime import datetime
from .utils.inventory_helpers import get_reserved_quantity, get_active_loans
from datetime import datetime, timedelta
from .utils.utils import flash_message, send_email
import hashlib
from flask import make_response
import csv
from io import StringIO
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from flask import send_file
import io
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from io import BytesIO, TextIOWrapper
import tempfile
import pandas as pd
from .utils.utils import flash_message, send_email
from app.services.reminder_service import ReminderService
from app.models import ReminderNotification
from datetime import timedelta
main_blueprint = Blueprint('main', __name__)

@main_blueprint.route('/')
@login_required
def index():
    messages = get_flashed_messages(with_categories=True)
    return redirect(url_for('main.dashboard'))

@main_blueprint.route('/logout')
@login_required  # Assicurati che solo gli utenti loggati possano accedere a questa route
def logout():
    logout_user()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('main.login'))  # Reindirizza l'utente alla pagina di login dopo il logout

@main_blueprint.route('/addUser', methods=['POST'])
def register_user():
    name = request.form.get('name')
    surname = request.form.get('surname')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')
    # Verifica che il ruolo sia uno di quelli permessi
    allowed_roles = {'Professor', 'RTD', 'Researcher', 'PhDStudent', 'Scholar'}
    if role not in allowed_roles:
        flash('Invalid role selected.','error')
        return redirect(url_for('main.show_register_page'))  # Assicurati di avere una route e una funzione per 'register_page'
    # Controlla che l'email non sia già utilizzata
    if User.query.filter_by(email=email).first() is not None:
        flash('Email already used.','error')
        return redirect(url_for('main.show_register_page'))
    # Crea un nuovo utente
    hashed_password = generate_password_hash(password)
    new_user = User(name=name, surname=surname, email=email, password_hash=hashed_password, role=role)
    db.session.add(new_user)
    db.session.commit()
    flash('User successfully registered.','success')
    return redirect(url_for('main.login'))



@main_blueprint.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash('Hello ' + user.name, category="success")
            return redirect(url_for('main.index'))
        else:
            flash('Invalid email or password.', 'error')
            return redirect(url_for('main.login'))
    return render_template('backend/auth-sign-in.html', messages=get_flashed_messages(with_categories=True))


@main_blueprint.route('/register')
def show_register_page():
    messages = get_flashed_messages(with_categories=True)
    return render_template('backend/register.html',messages=messages)

@main_blueprint.route('/lost_password')
def lost_password():
    messages = get_flashed_messages(with_categories=True)
    return render_template('backend/lost_password.html',messages=messages)

@main_blueprint.route('/add_product', methods=['POST', 'GET'])
@login_required
def add_product():
    if request.method == 'POST':
        name = request.form['name']
        unique_code = request.form['unique_code']
        description = request.form.get('description', '')
        location_id = request.form['location_id']
        project_id = request.form['project_id']
        category_ids = request.form.getlist('category_ids')
        
        # Validazione campi obbligatori
        if not name or not unique_code or not location_id or not project_id:
            flash('Please fill in all required fields.')
            return redirect(url_for('main.add_product'))
        
        # Gestione upload immagine
        image_path = None
        if 'product_image' in request.files:
            file = request.files['product_image']
            if file and file.filename != '':
                # Verifica il tipo di file
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
                if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                    # Crea nome file sicuro
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                    filename = secure_filename(f"product_{current_user.id}_{timestamp}.{file.filename.rsplit('.', 1)[1].lower()}")
                    
                    # Crea cartella se non esiste
                    upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'products')
                    if not os.path.exists(upload_path):
                        os.makedirs(upload_path)
                    
                    # Salva file
                    file_path = os.path.join(upload_path, filename)
                    file.save(file_path)
                    image_path = f"/static/uploads/products/{filename}"

        existing_product = Product.query.filter_by(unique_code=unique_code).first()
        if existing_product:
            flash('A product with this code already exists.')
            return redirect(url_for('main.add_product'))
        
        if not category_ids:
            flash('You must select at least one category.')
            return redirect(url_for('add_product_page'))
        
        # Verifica l'esistenza delle categorie
        valid_categories = Category.query.filter(Category.id.in_(category_ids)).count()
        if valid_categories != len(category_ids):
            flash('One or more selected categories are invalid.')
            return redirect(url_for('add_product_page'))
        categories = Category.query.filter(Category.id.in_(category_ids)).all()
        quantity = int(request.form['quantity'])
        location = Location.query.get(location_id)
        project = Project.query.get(project_id)
        owner_id = current_user.id
        product = Product(
            name=name,
            unique_code=unique_code,
            description=description,
            location_id=location.id,
            project_id=project.id,
            quantity=quantity,
            owner_id=owner_id,
            categories=categories,
            image_path=image_path
        )
        
        db.session.add(product)
        db.session.commit()
        # Aggiungi il proprietario come manager
        product_manager = ProductManager(product_id=product.id, user_id=owner_id)
        db.session.add(product_manager)
        db.session.commit()

        flash('Product added successfully!', 'success')
        return redirect(url_for('main.list_products'))

    categories = Category.query.all()
    projects = Project.query.all()
    locations = Location.query.all()
    messages = get_flashed_messages(with_categories=True)
    return render_template('backend/page-add-product.html', 
                         categories=categories, 
                         locations=locations, 
                         projects=projects, 
                         messages=messages)

@main_blueprint.route('/upload_csv', methods=['POST'])
@login_required
def upload_csv():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)

        with open(file_path, mode='r') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                # Assuming that the CSV has columns that exactly match your product model's required fields
                product = Product(
                    name=row['name'],
                    unique_code=row['unique_code'],
                    description=row['description'],
                    category_id=int(row['category_id']),
                    quantity=int(row['quantity'])
                )
                db.session.add(product)
            db.session.commit()

        flash('Products uploaded successfully!', 'success')
        return redirect(url_for('main.page_list_product'))
    return render_template('backend/upload_csv.html')

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ['csv']


@main_blueprint.route('/add_category', methods=['POST'])
@login_required
def add_category():
    name = request.form['name'].strip().lower()

    # Controlla se la categoria esiste già (case-insensitive)
    existing_category = Category.query.filter(func.lower(Category.name) == name).first()
    if existing_category:
        return jsonify({
            'status': 'error',
            'message': 'Category with this name already exists'
        }), 400

    new_category = Category(name=name)
    db.session.add(new_category)
    try:
        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': 'Category added successfully!',
            'entity': {'id': new_category.id, 'name': new_category.name}
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': 'Error adding category',
            'error': str(e)
        }), 400
        


@main_blueprint.route('/add_location', methods=['POST'])
@login_required
def add_location():
    pavilion = request.form.get('pavilion').strip().lower()
    room = request.form.get('room', '').strip().lower()
    cabinet = request.form.get('cabinet', '').strip().lower()
    
    if not pavilion:
        return jsonify({'error': 'Pavilion is required'}), 400

    # Controlla se la location esiste già (case-insensitive)
    existing_location = Location.query.filter_by(pavilion=pavilion, room=room, cabinet=cabinet).first()
    if existing_location:
        return jsonify({
            'status': 'error',
            'message': 'Location with this pavilion, room, and cabinet already exists'
        }), 400

    location = Location(pavilion=pavilion, room=room, cabinet=cabinet)
    db.session.add(location)
    try:
        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': 'Location added successfully',
            'entity': {'id': location.id, 'name': f"{location.pavilion} - {location.room} - {location.cabinet}"}
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': 'Error adding location',
            'error': str(e)
        }), 400

@main_blueprint.route('/add_project', methods=['POST']) #TODO non mi piace la lista. Ci vuole un  tasto di ricerca 
@login_required
def add_project():
    name = request.form.get('name').strip().lower()
    funding_body = request.form.get('funding_body', '').strip()

    if not name:
        return jsonify({'error': 'Project name is required'}), 400

    # Controlla se il progetto esiste già (case-insensitive)
    existing_project = Project.query.filter_by(name=name).first()
    if existing_project:
        return jsonify({
            'status': 'error',
            'message': 'Project with this name already exists'
        }), 400

    project = Project(name=name, funding_body=funding_body)
    db.session.add(project)
    try:
        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': 'Project added successfully',
            'entity': {'id': project.id, 'name': project.name}
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': 'Error adding the project',
            'error': str(e)
        }), 400

@main_blueprint.route('/list_products')
@login_required
def list_products():
    user_only = request.args.get('user_only', 'false').lower() in ['true', '1', 't']
    if user_only:
        # Mostra solo i prodotti dell'utente e quelli per cui è manager
        products = Product.query.filter(
            (Product.owner_id == current_user.id) |
            (Product.managers.any(User.id == current_user.id))
        ).distinct().all()
    else:
        # Mostra tutti i prodotti sulla piattaforma
        products = Product.query.all()

    print("Session in list_products:", dict(session))
    flash_message = session.pop('flash_message', None)
    if flash_message:
        flash(flash_message[1], flash_message[0])
    
    messages = get_flashed_messages(with_categories=True)
    return render_template('backend/page-list-product.html', products=products, messages=messages)

@main_blueprint.route('/list_disabled_products')
@login_required
def list_disabled_products():
    return _render_product_list(disabled_only=True)

def _render_product_list(disabled_only=False):
    user_only = request.args.get('user_only', 'false').lower() in ['true', '1', 't']
    
    query = Product.query

    if user_only:
        # Mostra solo i prodotti dell'utente e quelli per cui è manager
        query = query.filter(
            or_(Product.owner_id == current_user.id,
                Product.managers.any(User.id == current_user.id))
        )

    if disabled_only:
        query = query.filter_by(is_active=False)
    
    products = query.distinct().all()

    print("Session in list_products:", dict(session))
    flash_message = session.pop('flash_message', None)
    if flash_message:
        flash(flash_message[1], flash_message[0])
    
    messages = get_flashed_messages(with_categories=True)
    
    template = 'backend/page-list-product.html'
    return render_template(template, products=products, messages=messages, disabled_only=disabled_only)

@main_blueprint.route('/toggle_product/<int:product_id>', methods=['GET'])
@login_required
def toggle_product(product_id):
    product = Product.query.get_or_404(product_id)

    # Controlla se l'utente è il proprietario o un manager del prodotto
    if product.owner_id != current_user.id and current_user not in product.managers:
        flash("Non hai i permessi per modificare questo prodotto.", "danger")
        return redirect(url_for('main.list_products'))

    # Toggle is_active
    product.is_active = not product.is_active
    db.session.commit()

    flash("Lo stato del prodotto è stato aggiornato.", "success")
    return redirect(url_for('main.list_products'))

@main_blueprint.route('/product/<encrypted_id>')
@login_required
def view_product(encrypted_id):
    try:
        product_id = current_app.auth_s.loads(encrypted_id)  # Decodifica l'ID
    except Exception as e:
        flash('Invalid product ID.', 'error')
        return redirect(url_for('main.index'))

    product = Product.query.get_or_404(product_id)
    is_owner_or_manager = current_user.id == product.owner_id or any(manager.user_id == current_user.id for manager in product.manager_associations)
    all_categories = Category.query.order_by(Category.name).all()
    all_projects = Project.query.all()
    all_locations = Location.query.order_by(Location.pavilion).all()
    all_users = User.query.order_by(User.name).all()
    selected_categories = [category.id for category in product.categories]
    selected_project = product.project_id if product.project else None
    assigned_manager_ids = {manager.user_id for manager in product.manager_associations}
    reserved_dates = get_reserved_dates(product.id)
    reserved_quantity = get_reserved_quantity(product.id, datetime.now())
    
    messages = get_flashed_messages(with_categories=True)

    return render_template('backend/page-product.html',
                           encrypted_product_id=encrypted_id,
                           product=product,
                           all_categories=all_categories,
                           selected_categories=selected_categories,
                           all_users=all_users,
                           selected_managers=assigned_manager_ids,
                           all_projects=all_projects,
                           selected_project=selected_project,
                           all_locations=all_locations,
                           selected_location=product.location_id,
                           is_owner_or_manager=is_owner_or_manager,
                           reserved_dates=reserved_dates,
                           reserved_quantity=reserved_quantity,
                           messages=messages)

@main_blueprint.route('/approve_return/<int:loan_id>', methods=['POST'])
@login_required
def approve_return(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    
    # Verifica che il loan sia in stato 'in_review'
    if loan.status != 'in_review':
        return jsonify({'status': 'error', 'message': 'Only loans in review can be approved'}), 400
    
    # Verifica che l'utente sia il proprietario o un manager del prodotto
    if (loan.product.owner_id != current_user.id and 
        current_user not in loan.product.managers):
        return jsonify({'status': 'error', 'message': 'You are not authorized to perform this action'}), 403
    
    # Approva la restituzione - imposta lo stato a 'returned'
    loan.status = 'returned'
    db.session.commit()
    
    # Invia email di conferma al borrower
    try:
        send_email(
            subject=f"Return Confirmed - {loan.product.name}",
            recipient=loan.borrower.email,
            template='backend/return_approved_notification',
            borrower=loan.borrower,
            product=loan.product,
            loan=loan,
            manager=current_user,
            approval_date=datetime.now().strftime('%Y-%m-%d %H:%M')
        )
        print(f"Email sent to {loan.borrower.email} about approved return for {loan.product.name}")
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        # Non bloccare l'operazione se l'email fallisce
    
    return jsonify({'status': 'success', 'message': 'Return approved successfully. Borrower has been notified by email.'})
@main_blueprint.route('/reject_return/<int:loan_id>', methods=['POST'])
@login_required
def reject_return(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    
    # Verifica che il loan sia in stato 'in_review'
    if loan.status != 'in_review':
        return jsonify({'status': 'error', 'message': 'Only loans in review can be rejected'}), 400
    
    # Verifica che l'utente sia il proprietario o un manager del prodotto
    if (loan.product.owner_id != current_user.id and 
        current_user not in loan.product.managers):
        return jsonify({'status': 'error', 'message': 'You are not authorized to perform this action'}), 403
    
    # Rifiuta la restituzione - riporta lo stato a 'approved' 
    # (il materiale è ancora in prestito)
    loan.status = 'approved'
    db.session.commit()
    
    # Invia email di notifica al borrower
    try:
        send_email(
            subject=f"Return Rejected - {loan.product.name}",
            recipient=loan.borrower.email,
            template='backend/return_rejected_notification',
            borrower=loan.borrower,
            product=loan.product,
            loan=loan,
            manager=current_user,
            rejection_date=datetime.now().strftime('%Y-%m-%d %H:%M')
        )
        print(f"Email sent to {loan.borrower.email} about rejected return for {loan.product.name}")
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        # Non bloccare l'operazione se l'email fallisce
    
    return jsonify({'status': 'success', 'message': 'Return rejected - loan is still active. Borrower has been notified by email.'})



@main_blueprint.route('/set_product_unavailability/<encrypted_id>', methods=['POST'])
@login_required
def set_product_unavailability(encrypted_id):
    product_id = current_app.auth_s.loads(encrypted_id)
    product = Product.query.get_or_404(product_id)

    if current_user.id not in [product.owner_id] + [manager.user_id for manager in product.managers]:
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('main.index'))

    unavailability = Loan(
        product_id=product_id,
        borrower_id=current_user.id,
        manager_id=current_user.id,  # Considered as managed by self for unavailability
        start_date=request.form['start_date'],
        end_date=request.form['end_date'],
        status='unavailable',
        quantity=request.form['unavailability_quantity']
    )
    db.session.add(unavailability)
    try:
        db.session.commit()
        flash('Product set as unavailable successfully.', 'success')
    except:
        db.session.rollback()
        flash('Error setting product as unavailable.', 'error')

    return redirect(url_for('main.product_details', encrypted_id=encrypted_id))


@main_blueprint.route('/product_availability/<encrypted_id>')
@login_required
def product_availability(encrypted_id):
    try:
        product_id = current_app.auth_s.loads(encrypted_id)
    except Exception as e:
        return jsonify([])

    quantity_requested = int(request.args.get('quantity', 1))
    product = Product.query.get_or_404(product_id)
    loans = get_active_loans(product_id)

    events = []
    for loan in loans:
        if loan.quantity + quantity_requested > product.quantity:
            events.append({
                'title': 'Not Available',
                'start': loan.start_date.strftime('%Y-%m-%d'),
                'end': loan.end_date.strftime('%Y-%m-%d'),
                'color': 'red'
            })
        else:
            events.append({
                'title': 'Available',
                'start': loan.start_date.strftime('%Y-%m-%d'),
                'end': loan.end_date.strftime('%Y-%m-%d'),
                'color': 'green'
            })

    return jsonify(events)

@main_blueprint.route('/request_product/<encrypted_id>', methods=['POST'])
@login_required
def request_product(encrypted_id):
    try:
        product_id = current_app.auth_s.loads(encrypted_id)  # Decodifica l'ID
    except Exception as e:
        flash('Invalid product ID.', 'error')
        return redirect(url_for('main.index'))

    product = Product.query.get_or_404(product_id)

    request_quantity = int(request.form['request_quantity'])
    request_start_date = datetime.strptime(request.form['request_start_date'], '%Y-%m-%d').date()
    request_end_date = datetime.strptime(request.form['request_end_date'], '%Y-%m-%d').date()

    # Calcola le prenotazioni attive per questo prodotto
    active_loans = get_active_loans(product_id)

    # Logica per controllare la disponibilità del prodotto
    product_available = True
    for loan in active_loans:
        # Verifica se ci sono sovrapposizioni con le date di prenotazione richieste
        if (request_start_date <= loan.end_date and request_end_date >= loan.start_date):
            product_available = False
            break

    if product_available:
        # Salva la richiesta nel database o fai altro
        # Esempio: crea una nuova prenotazione nel database
        new_loan = Loan(
            product_id=product_id,
            borrower_id=current_user.id,
            manager_id=product.owner_id,
            start_date=request_start_date,
            end_date=request_end_date,
            quantity=request_quantity,
            status='pending'  # Può essere 'pending' o 'approved' a seconda della tua logica
        )
        db.session.add(new_loan)
        db.session.commit()

        flash('Product requested successfully!', 'success')
    else:
        flash('Product not available for the requested dates.', 'error')

    return redirect(url_for('main.view_product', encrypted_id=encrypted_id))

 


@main_blueprint.route('/fetch_availability/<encrypted_id>', methods=['GET'])
@login_required
def fetch_availability(encrypted_id):
    try:
        product_id = current_app.auth_s.loads(encrypted_id)
    except Exception as e:
        return jsonify([]), 400

    start_date = request.args.get('start')
    end_date = request.args.get('end')
    quantity = int(request.args.get('quantity', 1))

    if not start_date or not end_date:
        return jsonify([]), 400

    availability = get_availability(product_id, start_date, end_date, quantity)

    return jsonify(availability)

def get_availability(product_id, start_date, end_date, quantity):
    # Implementa la logica per ottenere la disponibilità del prodotto
    # Per esempio:
    product = Product.query.get(product_id)
    if not product:
        return []

    availability = []
    current_date = datetime.strptime(start_date, '%Y-%m-%d')
    end_date = datetime.strptime(end_date, '%Y-%m-%d')

    while current_date <= end_date:
        reserved_quantity = get_reserved_quantity(product_id, current_date)
        available = product.quantity - reserved_quantity >= quantity
        availability.append({'date': current_date.strftime('%Y-%m-%d'), 'available': available})
        current_date += timedelta(days=1)

    return availability

def get_reserved_quantity(product_id, date):
    reserved = db.session.query(func.sum(Loan.quantity)).filter(
        Loan.product_id == product_id,
        Loan.start_date <= date,
        Loan.end_date >= date
    ).scalar()

    return reserved if reserved else 0

def get_reserved_quantity_for_period(product_id, start_date, end_date):
    reserved_quantity = db.session.query(func.sum(Loan.quantity)).filter(
        Loan.product_id == product_id,
        Loan.start_date <= end_date,
        Loan.end_date >= start_date,
        Loan.status != 'returned'
    ).scalar()
    return reserved_quantity if reserved_quantity else 0

def get_reserved_dates(product_id):
    loans = Loan.query.filter_by(product_id=product_id).all()
    reserved_dates = {}
    for loan in loans:
        for date in daterange(loan.start_date, loan.end_date):
            reserved_dates[date.strftime('%Y-%m-%d')] = reserved_dates.get(date.strftime('%Y-%m-%d'), 0) + loan.quantity
    return reserved_dates

@main_blueprint.route('/process_booking/<encrypted_id>', methods=['POST'])
@login_required
def process_booking(encrypted_id):
    try:
        product_id = current_app.auth_s.loads(encrypted_id)
    except Exception as e:
        return jsonify({'status': 'error', 'message': 'Invalid product ID'}), 400

    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d')
    end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d')
    quantity = int(request.form['quantity'])

    if not check_date_range_availability(product_id, start_date, end_date, quantity):
        return jsonify({'status': 'error', 'message': 'Not enough quantity available for the selected dates'}), 400

    product = Product.query.get(product_id)

    # Verifica se l'utente loggato è manager o owner del prodotto
    if current_user.id == product.owner_id or current_user in product.managers:
        status = 'unavailable'
    else:
        status = 'pending'

    new_loan = Loan(
        product_id=product_id,
        borrower_id=current_user.id,
        manager_id=current_user.id,
        start_date=start_date,
        end_date=end_date,
        quantity=quantity,
        status=status
    )
    db.session.add(new_loan)
    db.session.commit()

    return jsonify({'status': 'success', 'message': 'Booking successful'})

def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + timedelta(n)

def check_date_range_availability(product_id, start_date, end_date, quantity):
    date = start_date
    while date <= end_date:
        reserved_quantity = get_reserved_quantity_for_day(product_id, date)
        product = Product.query.get(product_id)
        if product.quantity - reserved_quantity < quantity:
            return False
        date += timedelta(days=1)
    return True

@main_blueprint.route('/check_availability', methods=['POST'])
def check_availability():
    date = request.form['date']
    quantity = int(request.form['quantity'])
    product_id = int(request.form['product_id'])

    date = datetime.strptime(date, '%Y-%m-%d')

    reserved_quantity = get_reserved_quantity_for_day(product_id, date)
    product = Product.query.get(product_id)
    available_quantity = product.quantity - reserved_quantity

    return jsonify({'available': available_quantity >= quantity})

def get_reserved_quantity_for_day(product_id, date):
    reserved_quantity = db.session.query(func.sum(Loan.quantity)).filter(
        Loan.product_id == product_id,
        Loan.start_date <= date,
        Loan.end_date >= date,
        Loan.status != 'returned'
    ).scalar()
    return reserved_quantity if reserved_quantity else 0

@main_blueprint.route('/check_availability_range', methods=['POST'])
@login_required
def check_availability_range():
    try:
        product_id = current_app.auth_s.loads(request.form['encrypted_id'])
    except Exception as e:
        return jsonify({'status': 'error', 'message': 'Invalid product ID'}), 400

    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d')
    end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d')
    quantity = int(request.form['quantity'])

    reserved_dates = {}
    date = start_date
    while date <= end_date:
        reserved_quantity = get_reserved_quantity_for_day(product_id, date)
        product = Product.query.get(product_id)
        available_quantity = product.quantity - reserved_quantity
        reserved_dates[date.strftime('%Y-%m-%d')] = {'available': available_quantity >= quantity}
        date += timedelta(days=1)

    return jsonify(reserved_dates)

@main_blueprint.route('/loan_requests')
@login_required
def loan_requests():
    status = request.args.get('status', 'all')
    
    # Get outgoing loan requests
    outgoing_requests_query = Loan.query.filter_by(borrower_id=current_user.id)
    
    if status == 'in_review_pending':
        outgoing_requests_query = outgoing_requests_query.filter(Loan.status.in_(['in_review', 'pending']))
    elif status != 'all':
        outgoing_requests_query = outgoing_requests_query.filter(Loan.status == status)
    
    outgoing_requests = outgoing_requests_query.all()
    return render_template('backend/page-list-requests.html', outgoing_requests=outgoing_requests)


@main_blueprint.route('/get_loan_details/<int:loan_id>', methods=['GET'])
@login_required
def get_loan_details(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    
    # SOSTITUISCI QUESTA RIGA:
    # if loan.borrower_id != current_user.id and loan.manager_id != current_user.id:
    
    # CON QUESTA (controllo più ampio):
    if (loan.borrower_id != current_user.id and 
        loan.manager_id != current_user.id and
        loan.product.owner_id != current_user.id and 
        current_user not in loan.product.managers):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

    loan_details = {
        'product_name': loan.product.name,
        'quantity': loan.quantity,
        'start_date': loan.start_date.strftime('%Y-%m-%d'),
        'end_date': loan.end_date.strftime('%Y-%m-%d'),
        'location': f"{loan.product.location.pavilion}, Room: {loan.product.location.room}, Cabinet: {loan.product.location.cabinet}",
        'description': loan.product.description,
        'managers': [{'name': manager.name, 'surname': manager.surname, 'email': manager.email} for manager in loan.product.managers]
    }
    return jsonify(loan_details)

@main_blueprint.route('/restore_unavailable/<int:loan_id>', methods=['POST'])
@login_required
def restore_unavailable(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    if loan.status != 'unavailable':
        return jsonify({'status': 'error', 'message': 'Loan is not marked as unavailable'}), 400
    
    if current_user.id != loan.manager_id:
        return jsonify({'status': 'error', 'message': 'You are not authorized to perform this action'}), 403

    db.session.delete(loan)
    db.session.commit()
    
    return jsonify({'status': 'success', 'message': 'Product restored successfully'})

@main_blueprint.route('/cancel_request/<int:loan_id>', methods=['POST'])
@login_required
def cancel_request(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    if loan.status != 'pending':
        return jsonify({'status': 'error', 'message': 'Only pending loans can be cancelled'}), 400

    if current_user.id != loan.borrower_id:
        return jsonify({'status': 'error', 'message': 'You are not authorized to perform this action'}), 403

    loan.status = 'cancelled'
    db.session.commit()

    return jsonify({'status': 'success', 'message': 'Loan request cancelled successfully'})
#### Rotta per Segnare un Prestito come "Returned"
@main_blueprint.route('/mark_as_returned/<int:loan_id>', methods=['POST'])
@login_required
def mark_as_returned(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    if loan.status != 'approved':
        return jsonify({'status': 'error', 'message': 'Only approved loans can be marked as returned'}), 400

    if current_user.id != loan.borrower_id:
        return jsonify({'status': 'error', 'message': 'You are not authorized to perform this action'}), 403

    loan.status = 'in_review'
    db.session.commit()

    return jsonify({'status': 'success', 'message': 'Loan marked as returned, awaiting review'})

@main_blueprint.route('/requests_for_my_products')
@login_required
def requests_for_my_products():
    status = request.args.get('status', 'all')
    
    incoming_requests_query = Loan.query.join(Product).filter(
        (Product.owner_id == current_user.id) |
        (Product.managers.any(User.id == current_user.id))
    )
    
    if status != 'all':
        incoming_requests_query = incoming_requests_query.filter(Loan.status == status)

    incoming_requests = incoming_requests_query.all()

    unavailable_products = Loan.query.filter(
        (Loan.manager_id == current_user.id) &
        (Loan.status == 'unavailable')
    ).all()

    return render_template(
        'backend/requests_for_my_products.html',
        incoming_requests=incoming_requests,
        unavailable_products=unavailable_products
    )


@main_blueprint.route('/approve_request/<int:request_id>', methods=['POST'])
@login_required
def approve_request(request_id):
    request = Loan.query.get(request_id)
    if request and (request.product.owner_id == current_user.id or request.product.manager_id == current_user.id):
        request.status = 'approved'
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Request approved successfully'})
    return jsonify({'status': 'error', 'message': 'Request not found or not authorized'}), 404

@main_blueprint.route('/reject_request/<int:request_id>', methods=['POST'])
@login_required
def reject_request(request_id):
    request = Loan.query.get(request_id)
    if request and (request.product.owner_id == current_user.id or request.product.manager_id == current_user.id):
        request.status = 'rejected'
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Request rejected successfully'})
    return jsonify({'status': 'error', 'message': 'Request not found or not authorized'}), 404

@main_blueprint.route('/extend_loan/<int:loan_id>', methods=['POST'])
@login_required
def extend_loan(loan_id):
    loan = Loan.query.get_or_404(loan_id)
    new_end_date = request.form.get('new_end_date')

    if not new_end_date:
        return jsonify({'status': 'error', 'message': 'New termination date is required'}), 400

    try:
        new_end_date = datetime.strptime(new_end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Invalid date format'}), 400

    if loan.product.owner_id != current_user.id and not current_user in loan.product.managers:
        return jsonify({'status': 'error', 'message': 'You are not authorized to extend this loan'}), 403

    loan.end_date = new_end_date
    db.session.commit()

    return jsonify({'status': 'success', 'message': 'Loan extended successfully'})

@main_blueprint.route('/reset_password_request', methods=['GET', 'POST'])
def reset_password_request():
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        if user:
            token = generate_confirmation_token(user.email)
            send_email('Reset Your Password', user.email, '/backend/reset_password', token=token)
            flash('An email has been sent with instructions to reset your password.', 'info')
            return redirect(url_for('main.login'))
        else:
            flash('Email address not found.', 'warning')
    return redirect(url_for('main.lost_password'))

def generate_confirmation_token(email):
    serializer = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
    return serializer.dumps(email, salt=current_app.config['SECURITY_PASSWORD_SALT'])

@main_blueprint.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = confirm_token(token)
        if email is False:
            flash('The reset link is invalid or has expired.', 'danger')
            return redirect(url_for('main.reset_password_request'))
    except:
        flash('An error occurred. Please try again.', 'danger')
        return redirect(url_for('main.reset_password_request'))
    
    user = User.query.filter_by(email=email).first_or_404()

    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return render_template('/backend/reset_password.html', token=token)

        # Debug: Print the password (solo per scopi di debugging, rimuovi in produzione)
        print(f"Password inserita: {password}")

        # Debug: Genera un hash della password con un metodo noto
        debug_hash = hashlib.sha256(password.encode()).hexdigest()
        print(f"Debug hash: {debug_hash}")

        # Salva la password
        user.password = password

        # Debug: Stampa l'hash della password salvata
        print(f"Hash salvato: {user.password_hash}")

        try:
            db.session.commit()
            
            # Debug: Verifica se la password può essere verificata correttamente
            if check_password_hash(user.password_hash, password):
                print("La password può essere verificata correttamente.")
            else:
                print("ERRORE: La password non può essere verificata.")
                
        except Exception as e:
            print(f"Errore durante il salvataggio: {str(e)}")
            db.session.rollback()
        
        flash('Your password has been updated!', 'success')
        return redirect(url_for('main.login'))
    
    return render_template('/backend/reset_password.html', token=token)

def confirm_token(token, expiration=3600):
    serializer = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
    try:
        email = serializer.loads(
            token,
            salt=current_app.config['SECURITY_PASSWORD_SALT'],
            max_age=expiration
        )
    except SignatureExpired:
        return False  # valid token, but expired
    except BadSignature:
        return False  # invalid token
    return email


@main_blueprint.route('/users')
@login_required
def users():
    return render_template('/backend/page-list-users.html', users=users)

@main_blueprint.route('/apiusers')
@login_required
def api_users():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    per_page = 10

    if search:
        users_query = User.query.filter(
            (User.name.ilike(f'%{search}%')) |
            (User.surname.ilike(f'%{search}%')) |
            (User.role.ilike(f'%{search}%'))
        )
    else:
        users_query = User.query

    users_pagination = users_query.paginate(page=page, per_page=per_page)
    users = users_pagination.items

    users_list = []
    for user in users:
        # Calcola statistiche per ogni utente
        total_loans = Loan.query.filter_by(borrower_id=user.id).count()
        owned_products = Product.query.filter_by(owner_id=user.id).count()
        managed_products = ProductManager.query.filter_by(user_id=user.id).count()
        
        users_list.append({
            'id': user.id,
            'name': user.name,
            'surname': user.surname,
            'email': user.email,
            'role': user.role,
            'profile_image': user.profile_image or '/static/images/Blue_-_Squared_-_Delete_icon_(Wikiproject_icons).svg.png',
            'total_loans': total_loans,
            'owned_products': owned_products,
            'managed_products': managed_products
        })

    return jsonify({
        'users': users_list,
        'pagination': {
            'page': users_pagination.page,
            'pages': users_pagination.pages,
            'total': users_pagination.total,
            'prev_num': users_pagination.prev_num,
            'next_num': users_pagination.next_num,
            'has_prev': users_pagination.has_prev,
            'has_next': users_pagination.has_next,
        }
    })

@main_blueprint.route('/editProfile')
@login_required
def userprof():
    # Calcola le statistiche dell'utente
    total_loans = Loan.query.filter_by(borrower_id=current_user.id).count()
    owned_products = Product.query.filter_by(owner_id=current_user.id).count()
    managed_products = ProductManager.query.filter_by(user_id=current_user.id).count()
    
    user_stats = {
        'total_loans': total_loans,
        'owned_products': owned_products,
        'managed_products': managed_products
    }
    
    messages = get_flashed_messages(with_categories=True)
    return render_template('/backend/user-profile-edit.html', 
                         messages=messages, 
                         user_stats=user_stats)

@main_blueprint.route('/change_password_form', methods=['GET', 'POST'])
@login_required
def change_password_form():
    if request.method == 'POST':
        password = request.form.get('newpassword')
        confirm_password = request.form.get('validation')
        if password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return render_template('change_password.html')
        user = User.query.filter_by(email=current_user.email).first_or_404()
        if user:
            # Salva la nuova password
            user.password_hash = generate_password_hash(password)
            try:
                db.session.commit()
                
                # Debug: Verifica se la password può essere verificata correttamente
                if check_password_hash(user.password_hash, password):
                    print("La password può essere verificata correttamente.")
                else:
                    print("ERRORE: La password non può essere verificata.")
            except Exception as e:
                print(f"Errore durante il salvataggio: {str(e)}")
                db.session.rollback()
                flash('An error occurred while updating your password. Please try again.', 'danger')
                return render_template('change_password.html')

            flash('Your password has been updated!', 'success')
            return redirect(url_for('main.login'))
        else:
            flash('User not found.', 'danger')

    return render_template('change_password.html')



@main_blueprint.route('/download_excel_template', methods=['GET'])
@login_required
def download_excel_template():
    # Recupera i dati dal database
    categories = Category.query.all()
    projects = Project.query.all()
    locations = Location.query.all()
    managers = User.query.all()  # Aggiungi questa riga per recuperare gli utenti

    # Crea un nuovo workbook e attiva il foglio
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Template"

    # Definisci le intestazioni
    headers = ["Name", "Unique Code", "Description", "Pavilion", "Room", "Cabinet", "Project", "Categories", "Quantity", "Managers"]
    ws.append(headers)

    # Aggiungi un esempio
    example = ["Example Name", "Example Code", "Example Description", "Example Pavilion", "Example Room", "Example Cabinet", "Example Project", "Example Category", 1, "Example Manager"]
    ws.append(example)

    # Aggiungi una nuova scheda per i valori dei menu a tendina
    ws_data = wb.create_sheet(title="Data")
    ws_data.sheet_state = 'hidden'  # Nascondi la scheda

    # Aggiungi i dati di validazione
    category_list = [category.name for category in categories]
    project_list = [project.name for project in projects]
    pavilion_list = list(set([location.pavilion for location in locations]))
    room_list = list(set([location.room for location in locations if location.room]))
    cabinet_list = list(set([location.cabinet for location in locations if location.cabinet]))
    manager_list = [f"{user.name} {user.surname}" for user in managers]

    # Inserisci i dati nella scheda nascosta
    for idx, category in enumerate(category_list, start=1):
        ws_data[f'A{idx}'] = category
    for idx, project in enumerate(project_list, start=1):
        ws_data[f'B{idx}'] = project
    for idx, pavilion in enumerate(pavilion_list, start=1):
        ws_data[f'C{idx}'] = pavilion
    for idx, room in enumerate(room_list, start=1):
        ws_data[f'D{idx}'] = room
    for idx, cabinet in enumerate(cabinet_list, start=1):
        ws_data[f'E{idx}'] = cabinet
    for idx, manager in enumerate(manager_list, start=1):
        ws_data[f'F{idx}'] = manager

    # Definisci gli intervalli di nomi
    category_range = f'Data!$A$1:$A${len(category_list)}'
    project_range = f'Data!$B$1:$B${len(project_list)}'
    pavilion_range = f'Data!$C$1:$C${len(pavilion_list)}'
    room_range = f'Data!$D$1:$D${len(room_list)}'
    cabinet_range = f'Data!$E$1:$E${len(cabinet_list)}'
    manager_range = f'Data!$F$1:$F${len(manager_list)}'

    # Aggiungi la validazione dei dati
    dv_categories = DataValidation(type="list", formula1=category_range, allow_blank=True)
    dv_projects = DataValidation(type="list", formula1=project_range, allow_blank=True)
    dv_pavilions = DataValidation(type="list", formula1=pavilion_range, allow_blank=True)
    dv_rooms = DataValidation(type="list", formula1=room_range, allow_blank=True)
    dv_cabinets = DataValidation(type="list", formula1=cabinet_range, allow_blank=True)
    dv_managers = DataValidation(type="list", formula1=manager_range, allow_blank=True)

    ws.add_data_validation(dv_categories)
    ws.add_data_validation(dv_projects)
    ws.add_data_validation(dv_pavilions)
    ws.add_data_validation(dv_rooms)
    ws.add_data_validation(dv_cabinets)
    ws.add_data_validation(dv_managers)

    # Applica la validazione delle colonne appropriate
    max_row = ws.max_row + 1000  # aggiunge 1000 righe per validazione
    dv_categories.add(f'H2:H{max_row}')
    dv_projects.add(f'G2:G{max_row}')
    dv_pavilions.add(f'D2:D{max_row}')
    dv_rooms.add(f'E2:E{max_row}')
    dv_cabinets.add(f'F2:F{max_row}')
    dv_managers.add(f'J2:J{max_row}')

    # Salva il workbook in un oggetto BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='product_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@main_blueprint.route('/upload_excel', methods=['POST'])
@login_required
def upload_excel():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('main.add_product'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('main.add_product'))
    
    if file and file.filename.endswith(('.xlsx', '.xls')):
        filename = secure_filename(file.filename)
        
        # Crea la cartella 'temp' se non esiste
        temp_folder = os.path.join(current_app.root_path, 'temp')
        os.makedirs(temp_folder, exist_ok=True)
        
        file_path = os.path.join(temp_folder, filename)
        file.save(file_path)
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            products_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                name, code, description, pavilion, room, cabinet, project, categories, quantity, managers = row
                
                products_data.append({
                    'name': name,
                    'code': code,
                    'description': description,
                    'pavilion': pavilion,
                    'room': room,
                    'cabinet': cabinet,
                    'project': project,
                    'categories': categories.split(',') if categories else [],
                    'quantity': quantity,
                    'managers': managers.split(',') if managers else []
                })
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            return redirect(url_for('main.add_product'))
        finally:
            os.remove(file_path)  # Remove the temporary file

        # Serializza solo i dati necessari degli utenti
        all_users = [{'id': user.id, 'name': user.name, 'surname': user.surname} for user in User.query.all()]
        
        return render_template('backend/confirm_products.html', products=products_data, all_users=all_users)
    else:
        flash('Invalid file type. Please upload an Excel file.', 'error')
        return redirect(url_for('main.add_product'))
    
@main_blueprint.route('/process_products', methods=['POST'])
@login_required
def process_products():
    products_count = int(request.form.get('products_count', 0))
    error_occurred = False

    for i in range(products_count):
        name = request.form.get(f'name_{i}')
        code = request.form.get(f'code_{i}')
        description = request.form.get(f'description_{i}')
        pavilion = request.form.get(f'pavilion_{i}')
        room = request.form.get(f'room_{i}')
        cabinet = request.form.get(f'cabinet_{i}')
        project_name = request.form.get(f'project_{i}')
        categories = request.form.get(f'categories_{i}', '').split(',')
        quantity = request.form.get(f'quantity_{i}')
        manager_ids = request.form.getlist(f'managers_{i}[]')


        # Validazione dei campi obbligatori
        if not name or not code or not pavilion or not project_name or not quantity:
            flash(f'Product {i+1}: All fields except Description, Room, and Cabinet are required.', 'error')
            error_occurred = True
            continue

        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError
        except ValueError:
            flash(f'Product {i+1}: Quantity must be a positive integer.', 'error')
            error_occurred = True
            continue

        # Trova o crea la location
        location = Location.query.filter_by(pavilion=pavilion, room=room, cabinet=cabinet).first()
        if not location:
            location = Location(pavilion=pavilion, room=room, cabinet=cabinet)
            db.session.add(location)

        # Trova o crea il progetto
        project = Project.query.filter_by(name=project_name).first()
        if not project:
            project = Project(name=project_name)
            db.session.add(project)

        # Trova o crea le categorie
        category_objects = []
        for cat_name in categories:
            cat_name = cat_name.strip()
            if cat_name:
                category = Category.query.filter_by(name=cat_name).first()
                if not category:
                    category = Category(name=cat_name)
                    db.session.add(category)
                category_objects.append(category)

        # Crea il prodotto
        try:
            product = Product(
                name=name,
                unique_code=code,
                description=description,
                location=location,
                project=project,
                categories=category_objects,
                quantity=quantity,
                owner_id=current_user.id
            )
            db.session.add(product)
            db.session.flush()  # This will assign an ID to the product if it's valid

            # Aggiungi il proprietario (utente corrente) come manager
            owner_manager = ProductManager(product=product, user=current_user)
            db.session.add(owner_manager)

            # Aggiungi gli altri manager
            for manager_id in manager_ids:
                user = User.query.get(int(manager_id))
                if user and user != current_user:  # Evita di aggiungere il proprietario due volte
                    product_manager = ProductManager(product=product, user=user)
                    db.session.add(product_manager)

        except IntegrityError:
            db.session.rollback()
            flash(f'Product {i+1}: A product with this code already exists.', 'error')
            error_occurred = True
            continue

    if error_occurred:
        db.session.rollback()
        return redirect(url_for('main.add_product'))
    else:
        try:
            db.session.commit()
            flash('Products created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred while saving the products: {str(e)}', 'error')
            return redirect(url_for('main.add_product'))
    return redirect(url_for('main.list_products'))


@main_blueprint.route('/search_managers', methods=['GET'])
@login_required
def search_managers():
    term = request.args.get('term', '')
    managers = User.query.filter(
        (User.name.ilike(f'%{term}%') | User.surname.ilike(f'%{term}%')) &
        (User.id != current_user.id)
    ).all()
    return jsonify([{'id': m.id, 'name': m.name, 'surname': m.surname} for m in managers])



@main_blueprint.route('/export_products', methods=['POST'])
@login_required
def export_products():
    export_type = request.form.get('export_type')
    project_id = request.form.get('project_id')
    send_emails = request.form.get('send_emails') == 'true'

    if export_type == 'project':
        products = Product.query.filter_by(project_id=project_id).all()
    elif export_type == 'my_products':
        products = Product.query.filter_by(owner_id=current_user.id).all()
    else:
        flash('Invalid export type', 'error')
        return redirect(url_for('main.dashboard'))

    # Crea una lista di dizionari con i dati dei prodotti
    product_list = []
    for product in products:
        owner = User.query.get(product.owner_id)
        product_list.append({
            'Name': product.name,
            'Code': product.unique_code,
            'Description': product.description or '',
            'Pavilion': product.location.pavilion,
            'Room': product.location.room or '',
            'Cabinet': product.location.cabinet or '',
            'Project': product.project.name,
            'Categories': ', '.join([c.name for c in product.categories]),
            'Quantity': str(product.quantity),
            'Owner': f"{owner.name} {owner.surname}" if owner else "Unknown"
        })

    # Crea un DataFrame
    df = pd.DataFrame(product_list)

    # Salva il DataFrame in un file CSV temporaneo
    with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as temp_csv:
        csv_path = temp_csv.name
        df.to_csv(csv_path, index=False, encoding='utf-8')

    if send_emails and export_type == 'project':
        owners = set(User.query.get(product.owner_id) for product in products if product.owner_id != current_user.id)
        project = Project.query.get(project_id)
        for owner in owners:
            send_email(
                subject="Products Export Notification",
                recipient=owner.email,
                template='/backend/export_notification',
                user=owner,
                exporter=current_user,
                project=project
            )

    # Invia il file per il download
    return send_file(
        csv_path,
        mimetype='text/csv',
        download_name='products_export.csv',
        as_attachment=True
    )

    # Elimina il file temporaneo
    os.remove(csv_path)


@main_blueprint.route('/extract_page')
@login_required
def extract_page():
    projects = Project.query.all()
    return render_template('backend/extract_csv.html', projects=projects)


@main_blueprint.route('/dashboard')
@login_required
def dashboard():
    latest_products = Product.query.order_by(Product.id.desc()).limit(6).all()
    latest_loans = (
        Loan.query.filter_by(borrower_id=current_user.id)
        .order_by(Loan.start_date.desc())
        .limit(5)
        .all()
    )
    latest_returns = (
        Loan.query.filter_by(borrower_id=current_user.id)
        .filter(Loan.end_date != None)
        .order_by(Loan.end_date.desc())
        .limit(5)
        .all()
    )
    
    top_loaned_products = (
        db.session.query(Product, func.count(Loan.id).label('loans_count'))
        .join(Loan)
        .group_by(Product)
        .order_by(func.count(Loan.id).desc())
        .limit(10)
        .all()
    )
    # Statistiche per i commenti (aggiungi prima del return render_template)
    unread_comments_count = 0
    recent_comments = []
    total_comments_received = 0

    if current_user:
        # Commenti non letti sui propri prodotti
        unread_comments_count = Comment.query.join(Product).filter(
            (Product.owner_id == current_user.id) | 
            (Product.managers.any(User.id == current_user.id))
        ).filter(Comment.seen_by_manager == False).count()
        
        # Ultimi commenti ricevuti
        recent_comments = Comment.query.join(Product).filter(
            (Product.owner_id == current_user.id) | 
            (Product.managers.any(User.id == current_user.id))
        ).order_by(Comment.created_at.desc()).limit(3).all()
        
        # Totale commenti ricevuti sui propri prodotti
        total_comments_received = Comment.query.join(Product).filter(
            (Product.owner_id == current_user.id) | 
            (Product.managers.any(User.id == current_user.id))
        ).count()

        # Additional statistics
        total_products = Product.query.count()
        active_loans = Loan.query.filter(Loan.end_date == None).count()
        total_returns = Loan.query.filter(Loan.end_date != None).count()
        
        # Modifica qui: usiamo end_date invece di due_date
        overdue_loans = Loan.query.filter(Loan.end_date != None, Loan.end_date < datetime.utcnow()).count()

        messages = get_flashed_messages(with_categories=True)
        return render_template('backend/dashboard.html', 
                            latest_products=latest_products,
                            latest_loans=latest_loans,
                            latest_returns=latest_returns,
                            top_loaned_products=top_loaned_products,
                            total_products=total_products,
                            active_loans=active_loans,
                            total_returns=total_returns,
                            overdue_loans=overdue_loans,
                            messages=messages,
                            unread_comments_count=unread_comments_count,
                            recent_comments=recent_comments,
                            total_comments_received=total_comments_received)

@main_blueprint.route('/admin/reminders/status')
@login_required
def reminder_status():
    """Mostra lo stato del sistema promemoria (solo per admin)"""
    if current_user.role != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    reminder_service = ReminderService()
    
    # Statistiche
    total_notifications = db.session.query(ReminderNotification).count()
    recent_notifications = db.session.query(ReminderNotification).filter(
        ReminderNotification.sent_at >= datetime.utcnow() - timedelta(days=7)
    ).count()
    
    # Prestiti in scadenza
    loans_needing_reminders = reminder_service.get_loans_needing_reminders()
    
    context = {
        'config': reminder_service.config,
        'total_notifications': total_notifications,
        'recent_notifications': recent_notifications,
        'loans_needing_reminders': len(loans_needing_reminders),
        'is_enabled': reminder_service.config.get('settings', {}).get('enabled', False)
    }
    
    return render_template('backend/reminder_status.html', **context)

@main_blueprint.route('/admin/reminders/test', methods=['POST'])
@login_required
def test_reminders():
    """Testa l'invio di promemoria manualmente (solo per admin)"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        reminder_service = ReminderService()
        stats = reminder_service.process_all_reminders()
        
        return jsonify({
            'success': True,
            'message': f'Test completato. Processati: {stats["processed"]}, Inviati: {stats["sent"]}, Errori: {stats["errors"]}'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main_blueprint.route('/admin/reminders/config', methods=['GET', 'POST'])
@login_required
def reminder_config():
    """Gestisce la configurazione dei promemoria"""
    if current_user.role != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        # Salva la nuova configurazione (implementa secondo le tue esigenze)
        flash('Configurazione aggiornata', 'success')
        return redirect(url_for('main.reminder_status'))
    
    reminder_service = ReminderService()
    return render_template('backend/reminder_config.html', config=reminder_service.config)


@main_blueprint.route('/upload_avatar', methods=['POST'])
@login_required
def upload_avatar():
    try:
        if 'avatar' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        
        file = request.files['avatar']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        # Validazione del file
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'status': 'error', 'message': 'Invalid file type'}), 400
        
        # Crea un nome file sicuro
        filename = secure_filename(f"user_{current_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file.filename.rsplit('.', 1)[1].lower()}")
        
        # Crea la cartella se non esiste
        upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'avatars')
        if not os.path.exists(upload_path):
            os.makedirs(upload_path)
        
        # Salva il file
        file_path = os.path.join(upload_path, filename)
        file.save(file_path)
        
        # Rimuovi l'immagine precedente se esiste
        if current_user.profile_image and 'uploads' in current_user.profile_image:
            old_path = os.path.join(current_app.root_path, 'static', current_user.profile_image.lstrip('/static/'))
            if os.path.exists(old_path):
                os.remove(old_path)
        
        # Aggiorna il database
        current_user.profile_image = f"/static/uploads/avatars/{filename}"
        db.session.commit()
        
        return jsonify({
            'status': 'success', 
            'message': 'Profile picture updated successfully',
            'avatar_url': current_user.profile_image
        })
        
    except Exception as e:
        print(f"Error uploading avatar: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Error uploading image'}), 500

@main_blueprint.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    try:
        role = request.form.get('role')
        
        # Validazione del ruolo
        allowed_roles = {'Professor', 'RTD', 'Researcher', 'PhDStudent', 'Scholar'}
        if role not in allowed_roles:
            return jsonify({'status': 'error', 'message': 'Invalid role selected'}), 400
        
        # Aggiorna il ruolo
        current_user.role = role
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': 'Profile updated successfully'})
        
    except Exception as e:
        print(f"Error updating profile: {str(e)}")
        db.session.rollback()
        return jsonify({'status': 'error', 'message': 'An error occurred while updating profile'}), 500


@main_blueprint.route('/change_password', methods=['POST'])
@login_required
def change_password():
    try:
        current_password = request.form.get('currentPassword')
        new_password = request.form.get('newPassword')
        confirm_password = request.form.get('confirmPassword')
        
        # Validazioni
        if not current_password or not new_password or not confirm_password:
            return jsonify({'status': 'error', 'message': 'All fields are required'}), 400
        
        if new_password != confirm_password:
            return jsonify({'status': 'error', 'message': 'New passwords do not match'}), 400
        
        if len(new_password) < 8:
            return jsonify({'status': 'error', 'message': 'Password must be at least 8 characters long'}), 400
        
        # Verifica la password attuale
        if not check_password_hash(current_user.password_hash, current_password):
            return jsonify({'status': 'error', 'message': 'Current password is incorrect'}), 400
        
        # Aggiorna la password
        current_user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': 'Password changed successfully'})
        
    except Exception as e:
        print(f"Error changing password: {str(e)}")
        db.session.rollback()
        return jsonify({'status': 'error', 'message': 'An error occurred while changing password'}), 500



@main_blueprint.route('/get_user_activity')
@login_required
def get_user_activity():
    try:
        activities = []
        
        # Prestiti recenti dell'utente
        recent_loans = Loan.query.filter_by(borrower_id=current_user.id)\
                                .order_by(Loan.start_date.desc())\
                                .limit(10).all()
        
        for loan in recent_loans:
            time_diff = datetime.now() - loan.start_date
            if time_diff.days == 0:
                time_str = f"{max(1, time_diff.seconds // 3600)} hours ago"
            elif time_diff.days == 1:
                time_str = "1 day ago"
            else:
                time_str = f"{time_diff.days} days ago"
            
            activities.append({
                'icon': 'fas fa-handshake text-info',
                'title': 'Loan Request',
                'description': f'Requested {loan.product.name}',
                'time': time_str,
                'date': loan.start_date
            })
        
        # Ordina per data
        activities.sort(key=lambda x: x['date'], reverse=True)
        
        # Genera HTML
        html = ""
        for activity in activities[:5]:  # Solo i primi 5
            html += f"""
            <div class="d-flex mb-3">
                <i class="{activity['icon']} mt-1 mr-3"></i>
                <div>
                    <h6 class="mb-1">{activity['title']}</h6>
                    <p class="text-muted mb-0">{activity['description']}</p>
                    <small class="text-muted">{activity['time']}</small>
                </div>
            </div>
            """
        
        if not html:
            html = '<p class="text-muted text-center">No recent activity found</p>'
        
        return jsonify({'status': 'success', 'html': html})
        
    except Exception as e:
        print(f"Error getting activity: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Failed to load activity'}), 500
    

# Aggiungi queste routes al file app/views.py

@main_blueprint.route('/add_multiple_products')
@login_required
def add_multiple_products():
    """Pagina per inserire più prodotti manualmente"""
    categories = Category.query.all()
    projects = Project.query.all()
    locations = Location.query.all()
    all_users = User.query.all()
    
    messages = get_flashed_messages(with_categories=True)
    return render_template('backend/page-add-multiple-products.html', 
                         categories=categories, 
                         locations=locations, 
                         projects=projects,
                         all_users=all_users,
                         messages=messages)

@main_blueprint.route('/upload_product_image', methods=['POST'])
@login_required
def upload_product_image():
    """Upload immagine per prodotto (AJAX)"""
    try:
        if 'image' not in request.files:
            return jsonify({'status': 'error', 'message': 'No image file'}), 400
        
        file = request.files['image']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        # Verifica il tipo di file
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        if not (file and '.' in file.filename and 
                file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'status': 'error', 'message': 'Invalid file type'}), 400
        
        # Crea un nome file sicuro e unico
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        filename = secure_filename(f"product_{current_user.id}_{timestamp}.{file.filename.rsplit('.', 1)[1].lower()}")
        
        # Crea la cartella se non esiste
        upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'products')
        if not os.path.exists(upload_path):
            os.makedirs(upload_path)
        
        # Salva il file
        file_path = os.path.join(upload_path, filename)
        file.save(file_path)
        
        # Restituisci il path relativo
        relative_path = f"/static/uploads/products/{filename}"
        
        return jsonify({
            'status': 'success',
            'message': 'Image uploaded successfully',
            'image_path': relative_path,
            'filename': filename
        })
        
    except Exception as e:
        print(f"Error uploading product image: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Error uploading image'}), 500

@main_blueprint.route('/process_multiple_products', methods=['POST'])
@login_required
def process_multiple_products():
    """Elabora l'inserimento di più prodotti con immagini"""
    products_count = int(request.form.get('products_count', 0))
    error_occurred = False
    
    for i in range(products_count):
        name = request.form.get(f'name_{i}')
        code = request.form.get(f'code_{i}')
        description = request.form.get(f'description_{i}')
        pavilion = request.form.get(f'pavilion_{i}')
        room = request.form.get(f'room_{i}')
        cabinet = request.form.get(f'cabinet_{i}')
        project_name = request.form.get(f'project_{i}')
        categories = request.form.get(f'categories_{i}', '').split(',')
        quantity = request.form.get(f'quantity_{i}')
        manager_ids = request.form.getlist(f'managers_{i}[]')
        image_path = request.form.get(f'image_path_{i}')  # Nuovo campo per l'immagine

        # Validazione dei campi obbligatori
        if not name or not code or not pavilion or not project_name or not quantity:
            flash(f'Prodotto {i+1}: Tutti i campi eccetto Descrizione, Stanza e Armadio sono obbligatori.', 'error')
            error_occurred = True
            continue

        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError
        except ValueError:
            flash(f'Prodotto {i+1}: La quantità deve essere un numero positivo.', 'error')
            error_occurred = True
            continue

        # Trova o crea la location
        location = Location.query.filter_by(pavilion=pavilion, room=room, cabinet=cabinet).first()
        if not location:
            location = Location(pavilion=pavilion, room=room, cabinet=cabinet)
            db.session.add(location)

        # Trova o crea il progetto
        project = Project.query.filter_by(name=project_name).first()
        if not project:
            project = Project(name=project_name)
            db.session.add(project)

        # Trova o crea le categorie
        category_objects = []
        for cat_name in categories:
            cat_name = cat_name.strip()
            if cat_name:
                category = Category.query.filter_by(name=cat_name).first()
                if not category:
                    category = Category(name=cat_name)
                    db.session.add(category)
                category_objects.append(category)

        # Crea il prodotto con immagine
        try:
            product = Product(
                name=name,
                unique_code=code,
                description=description,
                location=location,
                project=project,
                categories=category_objects,
                quantity=quantity,
                owner_id=current_user.id,
                image_path=image_path  # Aggiungi il campo immagine
            )
            db.session.add(product)
            db.session.flush()

            # Aggiungi il proprietario come manager
            owner_manager = ProductManager(product=product, user=current_user)
            db.session.add(owner_manager)

            # Aggiungi gli altri manager
            for manager_id in manager_ids:
                user = User.query.get(int(manager_id))
                if user and user != current_user:
                    product_manager = ProductManager(product=product, user=user)
                    db.session.add(product_manager)

        except IntegrityError:
            db.session.rollback()
            flash(f'Prodotto {i+1}: Un prodotto con questo codice esiste già.', 'error')
            error_occurred = True
            continue

    if error_occurred:
        db.session.rollback()
        return redirect(url_for('main.add_multiple_products'))
    else:
        try:
            db.session.commit()
            flash(f'{products_count} prodotti creati con successo!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante il salvataggio dei prodotti: {str(e)}', 'error')
            return redirect(url_for('main.add_multiple_products'))
    
    return redirect(url_for('main.list_products'))

# Aggiungi o aggiorna questa route in views.py

@main_blueprint.route('/update_product/<encrypted_id>', methods=['POST'])
@login_required
def update_product(encrypted_id):
    try:
        product_id = current_app.auth_s.loads(encrypted_id)
        product = Product.query.get_or_404(product_id)
        
        # Verifica permessi
        if not (product.owner_id == current_user.id or 
                current_user in product.managers):
            flash('Non hai i permessi per modificare questo prodotto.', 'error')
            return redirect(url_for('main.list_products'))
        
        # Aggiorna campi base
        product.name = request.form.get('name', product.name)
        product.description = request.form.get('description', product.description)
        product.quantity = int(request.form.get('quantity', product.quantity))
        
        # Gestione upload nuova immagine
        if 'product_image' in request.files:
            file = request.files['product_image']
            if file and file.filename != '':
                # Verifica tipo file
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
                if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                    
                    # Rimuovi immagine precedente se esiste
                    if product.image_path and 'uploads' in product.image_path:
                        old_path = os.path.join(current_app.root_path, 'static', 
                                              product.image_path.lstrip('/static/'))
                        if os.path.exists(old_path):
                            os.remove(old_path)
                    
                    # Salva nuova immagine
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                    filename = secure_filename(f"product_{current_user.id}_{timestamp}.{file.filename.rsplit('.', 1)[1].lower()}")
                    
                    # Crea cartella se non esiste
                    upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'products')
                    if not os.path.exists(upload_path):
                        os.makedirs(upload_path)
                    
                    # Salva file
                    file_path = os.path.join(upload_path, filename)
                    file.save(file_path)
                    product.image_path = f"/static/uploads/products/{filename}"
                else:
                    flash('Tipo di file non supportato. Usa PNG, JPG, GIF o WebP.', 'error')
                    return redirect(url_for('main.view_product', encrypted_id=encrypted_id))
        
        # Gestione categorie (se incluso nel form)
        if 'category_ids' in request.form:
            category_ids = request.form.getlist('category_ids')
            if category_ids:
                categories = Category.query.filter(Category.id.in_(category_ids)).all()
                product.categories = categories
        
        # Gestione progetti (se incluso nel form)
        if 'project_id' in request.form:
            project_id = request.form.get('project_id')
            if project_id:
                product.project_id = int(project_id)
        
        # Gestione location (se incluso nel form)
        if 'location_id' in request.form:
            location_id = request.form.get('location_id')
            if location_id:
                product.location_id = int(location_id)
        
        # Gestione managers (se incluso nel form)
        if 'manager_ids' in request.form:
            manager_ids = request.form.getlist('manager_ids')
            if manager_ids:
                # Rimuovi manager esistenti (eccetto il proprietario)
                ProductManager.query.filter(
                    ProductManager.product_id == product.id,
                    ProductManager.user_id != product.owner_id
                ).delete()
                
                # Aggiungi nuovi manager
                for manager_id in manager_ids:
                    user = User.query.get(int(manager_id))
                    if user and user.id != product.owner_id:
                        product_manager = ProductManager(product_id=product.id, user_id=user.id)
                        db.session.add(product_manager)
        
        db.session.commit()
        flash('Prodotto aggiornato con successo!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante l\'aggiornamento: {str(e)}', 'error')
    
    return redirect(url_for('main.view_product', encrypted_id=encrypted_id))

# Route per rimuovere immagine prodotto
@main_blueprint.route('/remove_product_image/<encrypted_id>', methods=['POST'])
@login_required
def remove_product_image(encrypted_id):
    try:
        product_id = current_app.auth_s.loads(encrypted_id)
        product = Product.query.get_or_404(product_id)
        
        # Verifica permessi
        if not (product.owner_id == current_user.id or 
                current_user in product.managers):
            return jsonify({'status': 'error', 'message': 'Permessi insufficienti'}), 403
        
        # Rimuovi file fisico se esiste
        if product.image_path and 'uploads' in product.image_path:
            old_path = os.path.join(current_app.root_path, 'static', 
                                  product.image_path.lstrip('/static/'))
            if os.path.exists(old_path):
                os.remove(old_path)
        
        # Rimuovi da database
        product.image_path = None
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': 'Immagine rimossa con successo'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    


@main_blueprint.route('/upload_product_image/<encrypted_id>', methods=['POST'])
@login_required
def upload_product_image_update(encrypted_id):
    """Upload e aggiorna immagine per prodotto esistente"""
    print("=" * 50)
    print("INIZIO UPLOAD IMMAGINE PRODOTTO")
    print(f"encrypted_id ricevuto: {encrypted_id}")
    print("=" * 50)
    
    try:
        print("=== STEP 1: Decodifica ID ===")
        product_id = current_app.auth_s.loads(encrypted_id)
        print(f"product_id decodificato: {product_id}")
        
        print("=== STEP 2: Ricerca prodotto nel DB ===")
        product = Product.query.get_or_404(product_id)
        print(f"Prodotto trovato: {product.name} (ID: {product.id})")
        print(f"Image path attuale: {product.image_path}")
        
        print("=== STEP 3: Verifica permessi ===")
        print(f"current_user.id: {current_user.id}")
        print(f"product.owner_id: {product.owner_id}")
        # Verifica permessi
        if not (product.owner_id == current_user.id or 
                current_user in product.managers):
            print("=== ERRORE: Permessi insufficienti ===")
            return jsonify({'status': 'error', 'message': 'Permessi insufficienti'}), 403
        print("=== Permessi OK ===")
        
        print("=== STEP 4: Verifica files ricevuti ===")
        print(f"request.files.keys(): {list(request.files.keys())}")
        if 'image' not in request.files:
            print("=== ERRORE: Nessun file 'image' trovato ===")
            return jsonify({'status': 'error', 'message': 'No image file'}), 400
        
        file = request.files['image']
        print(f"File ricevuto: {file.filename}")
        print(f"File size: {file.content_length if hasattr(file, 'content_length') else 'Unknown'}")
        
        if file.filename == '':
            print("=== ERRORE: Nome file vuoto ===")
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        print("=== STEP 5: Validazione tipo file ===")
        # Verifica il tipo di file
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        if not (file and '.' in file.filename and 
                file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            print(f"=== ERRORE: Tipo file non valido: {file.filename} ===")
            return jsonify({'status': 'error', 'message': 'Invalid file type'}), 400
        print("=== Tipo file OK ===")
        
        print("=== STEP 6: Preparazione salvataggio ===")
        # Salva la vecchia immagine per rimuoverla dopo
        old_image_path = product.image_path
        print(f"Vecchia immagine: {old_image_path}")
        
        # Crea un nome file sicuro e unico
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        filename = secure_filename(f"product_{current_user.id}_{timestamp}.{file.filename.rsplit('.', 1)[1].lower()}")
        print(f"Nuovo filename: {filename}")
        
        # Crea la cartella se non esiste
        upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'products')
        print(f"Upload path: {upload_path}")
        if not os.path.exists(upload_path):
            print("=== Creazione cartella upload ===")
            os.makedirs(upload_path)
        else:
            print("=== Cartella upload già esistente ===")
        
        print("=== STEP 7: Salvataggio file fisico ===")
        # Salva il file
        file_path = os.path.join(upload_path, filename)
        print(f"File path completo: {file_path}")
        file.save(file_path)
        print("=== File salvato con successo ===")
        
        print("=== STEP 8: Aggiornamento database ===")
        # Aggiorna il database
        new_image_path = f"/static/uploads/products/{filename}"
        print(f"Nuovo image_path per DB: {new_image_path}")
        print(f"PRIMA - product.image_path: {product.image_path}")
        
        product.image_path = new_image_path
        print(f"DOPO assegnazione - product.image_path: {product.image_path}")
        
        print("=== Esecuzione commit ===")
        db.session.commit()
        print("=== COMMIT ESEGUITO CON SUCCESSO ===")
        
        # Verifica che sia stato salvato
        print("=== STEP 9: Verifica salvataggio ===")
        updated_product = Product.query.get(product_id)
        print(f"Prodotto ricaricato dal DB - image_path: {updated_product.image_path}")
        
        print("=== STEP 10: Rimozione vecchia immagine ===")
        # Rimuovi la vecchia immagine solo dopo il commit riuscito
        if old_image_path and 'uploads' in old_image_path:
            old_path = os.path.join(current_app.root_path, 'static', 
                                  old_image_path.lstrip('/static/'))
            print(f"Tentativo rimozione vecchia immagine: {old_path}")
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                    print("=== Vecchia immagine rimossa ===")
                except OSError as e:
                    print(f"=== Errore rimozione vecchia immagine: {e} ===")
            else:
                print("=== Vecchia immagine non trovata sul disco ===")
        else:
            print("=== Nessuna vecchia immagine da rimuovere ===")
        
        print("=== STEP 11: Risposta finale ===")
        response_data = {
            'status': 'success',
            'message': 'Image uploaded successfully',
            'image_path': new_image_path
        }
        print(f"Risposta: {response_data}")
        print("=" * 50)
        print("FINE UPLOAD IMMAGINE - SUCCESSO")
        print("=" * 50)
        
        return jsonify(response_data)
        
    except Exception as e:
        print("=" * 50)
        print(f"ERRORE DURANTE UPLOAD: {str(e)}")
        print(f"Tipo errore: {type(e).__name__}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        print("=" * 50)
        
        db.session.rollback()
        print("=== ROLLBACK ESEGUITO ===")
        
        # Se è stato salvato un nuovo file ma c'è stato un errore nel DB, rimuovilo
        if 'new_image_path' in locals():
            try:
                new_file_path = os.path.join(current_app.root_path, 'static', 
                                           new_image_path.lstrip('/static/'))
                if os.path.exists(new_file_path):
                    os.remove(new_file_path)
                    print("=== File nuovo rimosso dopo errore ===")
            except OSError:
                print("=== Errore nella rimozione file nuovo ===")
        
        return jsonify({'status': 'error', 'message': f'Error uploading image: {str(e)}'}), 500
    

@main_blueprint.route('/toggle_wishlist/<int:product_id>', methods=['POST'])
@login_required
def toggle_wishlist(product_id):
    product = Product.query.get_or_404(product_id)
    
    try:
        if product in current_user.wishlist:
            current_user.wishlist.remove(product)
            action = 'removed'
            message = f'Rimosso "{product.name}" dalla wishlist'
            icon = '💔'
        else:
            current_user.wishlist.append(product)
            action = 'added'
            message = f'Aggiunto "{product.name}" alla wishlist'
            icon = '❤️'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'action': action,
            'message': message,
            'icon': icon,
            'wishlist_count': len(current_user.wishlist)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'Errore durante l\'operazione'
        }), 500

@main_blueprint.route('/my_wishlist')
@login_required
def my_wishlist():
    """Pagina della wishlist dell'utente"""
    wishlist_products = current_user.wishlist
    all_users = User.query.all()
    
    # Cripta gli ID dei prodotti per i link
    for product in wishlist_products:
        product.encrypted_id = current_app.auth_s.dumps(product.id)
    
    messages = get_flashed_messages(with_categories=True)
    return render_template('backend/page-wishlist.html', 
                         products=wishlist_products, 
                         all_users=all_users,
                         messages=messages,
                         page_title="La Mia Wishlist")

@main_blueprint.route('/wishlist_count')
@login_required
def wishlist_count():
    """API per ottenere il conteggio wishlist (per badge nel menu)"""
    return jsonify({'count': len(current_user.wishlist)})

@main_blueprint.route('/mark_as_returned_with_comment/<int:loan_id>', methods=['POST'])
@login_required
def mark_as_returned_with_comment(loan_id):
    """Segna un prestito come ritornato con commento opzionale"""
    loan = Loan.query.get_or_404(loan_id)
    
    if loan.status != 'approved':
        return jsonify({'status': 'error', 'message': 'Only approved loans can be marked as returned'}), 400

    if current_user.id != loan.borrower_id:
        return jsonify({'status': 'error', 'message': 'You are not authorized to perform this action'}), 403

    # Ottieni i dati del commento dalla richiesta
    comment_text = request.form.get('comment_text', '').strip()
    rating = request.form.get('rating')
    
    # Segna il prestito come in review
    loan.status = 'in_review'
    
    # Se c'è un commento, crealo
    if comment_text:
        try:
            rating_value = int(rating) if rating and rating.isdigit() and 1 <= int(rating) <= 5 else None
        except (ValueError, TypeError):
            rating_value = None
            
        comment = Comment(
            loan_id=loan.id,
            borrower_id=current_user.id,
            product_id=loan.product_id,
            comment_text=comment_text,
            rating=rating_value
        )
        db.session.add(comment)
        loan.has_comment = True  # Aggiungi questo campo al modello Loan
    
    db.session.commit()
    
    # Invia email di notifica al manager includendo il commento se presente
    try:
        managers_emails = []
        if loan.product.owner:
            managers_emails.append(loan.product.owner.email)
        for manager in loan.product.managers:
            if manager.email not in managers_emails:
                managers_emails.append(manager.email)
        
        for manager_email in managers_emails:
            send_email(
                subject=f"Return Request - {loan.product.name}" + (" (with comment)" if comment_text else ""),
                recipient=manager_email,
                template='backend/return_request_notification',
                borrower=current_user,
                product=loan.product,
                loan=loan,
                comment=comment if comment_text else None,
                return_date=datetime.now().strftime('%Y-%m-%d %H:%M')
            )
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
    
    success_message = 'Loan marked as returned successfully'
    if comment_text:
        success_message += ' with your comment'
    success_message += ', awaiting manager review'
    
    return jsonify({'status': 'success', 'message': success_message})


@main_blueprint.route('/view_loan_comments/<int:loan_id>')
@login_required
def view_loan_comments(loan_id):
    """Visualizza i commenti di un prestito specifico"""
    loan = Loan.query.get_or_404(loan_id)
    
    # Verifica autorizzazione: borrower, owner del prodotto o manager
    if (current_user.id != loan.borrower_id and 
        current_user.id != loan.product.owner_id and 
        current_user not in loan.product.managers):
        flash('You are not authorized to view these comments.', 'error')
        return redirect(url_for('main.dashboard'))
    
    comments = Comment.query.filter_by(loan_id=loan_id).order_by(Comment.created_at.desc()).all()
    
    # Se l'utente è manager, segna i commenti come visti
    if (current_user.id == loan.product.owner_id or current_user in loan.product.managers):
        for comment in comments:
            if not comment.seen_by_manager:
                comment.seen_by_manager = True
        db.session.commit()
    
    # AGGIUNGI IL CALCOLO DELLE STATISTICHE
    total_comments = len(comments)
    rated_comments = [c for c in comments if c.rating is not None]
    average_rating = sum(c.rating for c in rated_comments) / len(rated_comments) if rated_comments else None
    total_ratings = len(rated_comments)
    
    return render_template('backend/loan_comments.html', 
                         loan=loan, 
                         comments=comments,
                         product=loan.product,
                         total_comments=total_comments,
                         average_rating=average_rating,
                         total_ratings=total_ratings)


@main_blueprint.route('/dashboard_comments_summary')
@login_required
def dashboard_comments_summary():
    """API endpoint per ottenere un riassunto dei commenti nel dashboard"""
    
    # Commenti non visti sui propri prodotti
    unread_comments = Comment.query.join(Product).filter(
        (Product.owner_id == current_user.id) | 
        (Product.managers.any(User.id == current_user.id))
    ).filter(Comment.seen_by_manager == False).count()
    
    # Ultimi commenti sui propri prodotti
    recent_comments = Comment.query.join(Product).filter(
        (Product.owner_id == current_user.id) | 
        (Product.managers.any(User.id == current_user.id))
    ).order_by(Comment.created_at.desc()).limit(5).all()
    
    return jsonify({
        'unread_count': unread_comments,
        'recent_comments': [{
            'id': c.id,
            'product_name': c.product.name,
            'borrower_name': f"{c.borrower.name} {c.borrower.surname}",
            'comment_text': c.comment_text[:100] + ('...' if len(c.comment_text) > 100 else ''),
            'rating': c.rating,
            'created_at': c.created_at.strftime('%d/%m/%Y %H:%M'),
            'seen': c.seen_by_manager
        } for c in recent_comments]
    })

@main_blueprint.route('/get_loan_comment_details/<int:loan_id>', methods=['GET'])
@login_required
def get_loan_comment_details(loan_id):
    """API per ottenere i dettagli del commento di un prestito"""
    loan = Loan.query.get_or_404(loan_id)
    
    # Verifica autorizzazioni (stesso controllo degli altri dettagli)
    if (loan.borrower_id != current_user.id and 
        loan.manager_id != current_user.id and
        loan.product.owner_id != current_user.id and 
        current_user not in loan.product.managers):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
    
    # Cerca il commento per questo prestito
    comment = Comment.query.filter_by(loan_id=loan_id).first()
    
    if not comment:
        return jsonify({'status': 'error', 'message': 'No comment found for this loan'}), 404
    
    # Prepara la risposta
    comment_data = {
        'id': comment.id,
        'comment_text': comment.comment_text,
        'rating': comment.rating,
        'created_at': comment.created_at.strftime('%d/%m/%Y alle %H:%M'),
        'borrower_name': f"{comment.borrower.name} {comment.borrower.surname}",
        'seen_by_manager': comment.seen_by_manager
    }
    
    # Segna il commento come visto se l'utente è un manager
    if (current_user.id == loan.product.owner_id or current_user in loan.product.managers):
        if not comment.seen_by_manager:
            comment.seen_by_manager = True
            db.session.commit()
    
    return jsonify({
        'status': 'success',
        'comment': comment_data,
        'product_name': loan.product.name,
        'product_id': loan.product_id,
        'loan_id': loan_id
    })


# OPZIONALE: Vista per ottenere tutti i commenti di un prodotto (se non l'hai già)
@main_blueprint.route('/product_comments/<int:product_id>')
@login_required  
def product_comments(product_id):
    """Visualizza tutti i commenti per un prodotto specifico"""
    product = Product.query.get_or_404(product_id)
    
    # Verifica autorizzazione: owner del prodotto o manager
    if (current_user.id != product.owner_id and current_user not in product.managers):
        flash('You are not authorized to view these comments.', 'error')
        return redirect(url_for('main.dashboard'))
    
    # Ottieni tutti i commenti per questo prodotto
    comments = Comment.query.filter_by(product_id=product_id).join(Loan).order_by(Comment.created_at.desc()).all()
    
    # Segna i commenti come visti se l'utente è manager
    if current_user.id == product.owner_id or current_user in product.managers:
        for comment in comments:
            if not comment.seen_by_manager:
                comment.seen_by_manager = True
        db.session.commit()
    
    # Calcola statistiche
    total_comments = len(comments)
    rated_comments = [c for c in comments if c.rating is not None]
    average_rating = sum(c.rating for c in rated_comments) / len(rated_comments) if rated_comments else None
    
    return render_template('backend/product_comments.html', 
                         product=product, 
                         comments=comments,
                         total_comments=total_comments,
                         average_rating=average_rating,
                         total_ratings=len(rated_comments))